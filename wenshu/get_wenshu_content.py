import base64
import json
from openpyxl import load_workbook
from Crypto.Cipher import DES3
from Crypto.Util.Padding import unpad

# 加载 Excel 文件
workbook = load_workbook("/Users/qsmy/Downloads/result_minshi.xlsx")

# 获取工作表
print(workbook.sheetnames)  # 打印所有工作表名称
read_sheet = workbook.worksheets[1]  # 通过索引获取
write_sheet = workbook.worksheets[2]  # 通过索引获取

# 初始化向量
iv = b"20250102"

# 写入表头
write_sheet.append(['s1', 's2', 's3', 's5', 's6', 's7', 's8', 's9',
                    's10', 's11', 's17', 's22', 's23', 's25', 's26',
                    's27', 's28', 's31', 's32', 's41', 's43', 's45',
                    's47', 's51', 'viewCount', 'wenshuAy'])

# 遍历所有行
for index, row in enumerate(read_sheet.iter_rows(values_only=True, min_row=2)):
    key = row[2]
    # 解密数据
    try:
        encrypted_data = base64.b64decode(row[4])
    except Exception as e:
        print(f"Error decoding row {index + 1}: {e}")
        write_sheet.append([])
        print(write_sheet)
        continue
    cipher_decrypt = DES3.new(key, DES3.MODE_CBC, iv)
    decrypted_data = cipher_decrypt.decrypt(encrypted_data)
    plaintext = unpad(decrypted_data, DES3.block_size).decode("utf-8")
    json_object = json.loads(plaintext)

    # 构建数据行
    data = [
        json_object.get('s1', ''), json_object.get('s2', ''), json_object.get('s3', ''),
        json_object.get('s5', ''), json_object.get('s6', ''), json_object.get('s7', ''),
        json_object.get('s8', ''), json_object.get('s9', ''), json_object.get('s10', ''),
        json_object.get('s11', ''), json_object.get('s17', ''), json_object.get('s22', ''),
        json_object.get('s23', ''), json_object.get('s25', ''), json_object.get('s26', ''),
        json_object.get('s27', ''), json_object.get('s28', ''), json_object.get('s31', ''),
        json_object.get('s32', ''), json_object.get('s41', ''), json_object.get('s43', ''),
        json_object.get('s45', ''), json_object.get('s47', ''), json_object.get('s51', ''),
        json_object.get('viewCount', ''), json_object.get('wenshuAy', '')
    ]

    # 将所有值转换为字符串
    data = [str(item) for item in data]

    # 写入数据
    write_sheet.append(data)
    print(write_sheet)

# 保存文件
output_path = "/Users/qsmy/Downloads/result_minshi.xlsx"
workbook.save(output_path)
print(f"File saved to {output_path}")