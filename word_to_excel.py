from docx import Document
from openpyxl import Workbook


def word_to_excel(docx_path, excel_path):
    # 打开Word文档
    doc = Document(docx_path)

    # 创建一个新的Excel工作簿
    wb = Workbook()
    ws = wb.active  # 选择活动的工作表

    # 遍历Word文档中的所有表格
    for table_idx, table in enumerate(doc.tables):
        # 可以为每个表格添加一个新的工作表，或者将所有表格写入同一个工作表
        # 这里选择将所有表格数据写入同一个工作表，并在每个表格前添加标题
        ws.append([f"Table {table_idx + 1}"])

        # 写入表头
        for col_idx, row in enumerate(table.rows):
            if col_idx == 0:
                # 假设第一行为表头
                headers = [cell.text.strip() for cell in row.cells]
                ws.append(headers)
            else:
                # 写入数据行
                data = [cell.text.strip() for cell in row.cells]
                ws.append(data)

        # 表格之间添加空行
        ws.append([])  # 空行

    # 删除最后一个多余的空白行（如果有）
    # if ws.max_row > 1:
    #     ws.delete_row(ws.max_row)

    # 保存Excel文件
    wb.save(excel_path)
    print(f"已成功将Word表格转换为Excel文件：{excel_path}")


# 使用示例
word_to_excel(
    "C:\\Users\\qsmy\\Desktop\\文书评分(1).docx",
    "C:\\Users\\qsmy\\Desktop\\output.xlsx",
)
